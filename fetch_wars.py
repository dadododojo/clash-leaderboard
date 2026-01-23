import requests
import json
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import os
from config import API_KEY, CLAN_TAG

# Try to import Discord webhook URL (optional)
try:
    from config import DISCORD_WEBHOOK_URL
except ImportError:
    DISCORD_WEBHOOK_URL = None

# API Configuration
BASE_URL = "https://api.clashofclans.com/v1"
HEADERS = {
    "Authorization": f"Bearer {API_KEY}",
    "Accept": "application/json"
}

EXCEL_FILE = 'clash_wars.xlsx'
LEADERBOARD_FILE = 'leaderboard.json'
ROSTER_SHEET_NAME = 'ROSTER'
MISSED_HITS_SHEET_NAME = 'MISSED_HITS'

def format_tag(tag):
    """Ensure tag starts with #"""
    if not tag.startswith('#'):
        return '#' + tag
    return tag

def fetch_current_war():
    """Fetch current war details"""
    clan_tag = format_tag(CLAN_TAG).replace('#', '%23')
    url = f"{BASE_URL}/clans/{clan_tag}/currentwar"
    
    try:
        response = requests.get(url, headers=HEADERS)
        response.raise_for_status()
        return response.json()
    except requests.exceptions.RequestException as e:
        print(f"Error fetching current war: {e}")
        return None

def get_war_id(war):
    """Generate unique war ID from preparation start time"""
    prep_time = war.get('preparationStartTime', '')
    if not prep_time:
        prep_time = war.get('endTime', str(datetime.now().timestamp()))
    return prep_time.replace(':', '-').replace('.', '-')

def is_war_ended(war):
    """Check if war has ended"""
    return war.get('state') == 'warEnded'

def is_cwl_war(war):
    """Check if this is a CWL war"""
    war_league = war.get('warLeague', {})
    return war_league.get('name') != 'Unranked' if war_league else False

def war_already_saved(war_id):
    """Check if war is already saved in Excel and if it's complete"""
    if not os.path.exists(EXCEL_FILE):
        return False, False
    
    try:
        book = load_workbook(EXCEL_FILE, read_only=True)
        sheet_name = war_id[:31]
        exists = sheet_name in book.sheetnames
        
        is_complete = False
        if exists:
            df = pd.read_excel(EXCEL_FILE, sheet_name=sheet_name)
            if not df.empty and 'War Complete' in df.columns:
                is_complete = df['War Complete'].iloc[0] == 'Yes'
        
        book.close()
        return exists, is_complete
    except Exception as e:
        print(f"Error checking existing wars: {e}")
        return False, False

def get_existing_loot_hits(war_id):
    """Get existing loot hit markings for a war if it exists"""
    if not os.path.exists(EXCEL_FILE):
        return {}
    
    try:
        sheet_name = war_id[:31]
        df = pd.read_excel(EXCEL_FILE, sheet_name=sheet_name)
        
        loot_hits = {}
        for idx, row in df.iterrows():
            player_tag = row['Player Tag']
            if 'Is Loot Hit' in row and row['Is Loot Hit'] in ['TRUE', True, 'Yes', 'yes', 1]:
                attack_num = row.get('Attack Number', 0)
                if player_tag not in loot_hits:
                    loot_hits[player_tag] = []
                loot_hits[player_tag].append(attack_num)
        
        return loot_hits
    except Exception as e:
        return {}

def process_war(war, preserve_loot_markings=True):
    """Process war data and extract player statistics"""
    if not war:
        return None
    
    war_id = get_war_id(war)
    existing_loot = get_existing_loot_hits(war_id) if preserve_loot_markings else {}
    
    war_state = war.get('state', 'unknown')
    end_time = war.get('endTime', 'N/A')
    team_size = war.get('teamSize', 0)
    is_complete = is_war_ended(war)
    
    clan = war.get('clan', {})
    members = clan.get('members', [])
    
    war_details = []
    
    for member in members:
        player_tag = member.get('tag', '')
        attacks = member.get('attacks', [])
        
        for attack_idx, attack in enumerate(attacks, 1):
            stars = attack.get('stars', 0)
            destruction = attack.get('destructionPercentage', 0)
            
            is_missed = (stars == 0 and destruction == 0)
            
            is_loot = False
            if player_tag in existing_loot and attack_idx in existing_loot[player_tag]:
                is_loot = True
            
            war_details.append({
                'War ID': war_id,
                'War State': war_state,
                'War Complete': 'Yes' if is_complete else 'No',
                'War End Time': end_time,
                'Team Size': team_size,
                'Player Name': member.get('name', ''),
                'Player Tag': player_tag,
                'Town Hall': member.get('townhallLevel', 0),
                'Map Position': member.get('mapPosition', 0),
                'Attack Number': attack_idx,
                'Stars': stars,
                'Destruction %': destruction,
                'Is Triple': 'Yes' if stars == 3 else 'No',
                'Is Missed': 'Yes' if is_missed else 'No',
                'Is Loot Hit': 'Yes' if is_loot else 'No'
            })
        
        if not attacks:
            war_details.append({
                'War ID': war_id,
                'War State': war_state,
                'War Complete': 'Yes' if is_complete else 'No',
                'War End Time': end_time,
                'Team Size': team_size,
                'Player Name': member.get('name', ''),
                'Player Tag': player_tag,
                'Town Hall': member.get('townhallLevel', 0),
                'Map Position': member.get('mapPosition', 0),
                'Attack Number': 0,
                'Stars': 0,
                'Destruction %': 0,
                'Is Triple': 'No',
                'Is Missed': 'Yes',
                'Is Loot Hit': 'No'
            })
    
    return {
        'war_id': war_id,
        'war_details': war_details,
        'is_ended': is_complete
    }

def save_war_to_excel(war_data):
    """Save war data to Excel with separate sheet for each war"""
    if not war_data or not war_data['war_details']:
        print("No war data to save")
        return False
    
    war_id = war_data['war_id']
    df = pd.DataFrame(war_data['war_details'])
    
    if os.path.exists(EXCEL_FILE):
        book = load_workbook(EXCEL_FILE)
    else:
        book = Workbook()
        if 'Sheet' in book.sheetnames:
            del book['Sheet']
    
    sheet_name = war_id[:31]
    
    if sheet_name in book.sheetnames:
        del book[sheet_name]
        print(f"Updating existing war {sheet_name}...")
    else:
        print(f"Creating new war {sheet_name}...")
    
    ws = book.create_sheet(sheet_name)
    
    for r_idx, r in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        ws.append(r)
        
        if r_idx == 1:
            for cell in ws[r_idx]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
    
    for column in ws.columns:
        max_length = 0
        column = list(column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width
    
    book.save(EXCEL_FILE)
    print(f"✓ Saved war to Excel")
    return True

def update_missed_hits_sheet():
    """Create/update a sheet tracking missed attacks"""
    if not os.path.exists(EXCEL_FILE):
        print("No war data file found")
        return
    
    book = load_workbook(EXCEL_FILE)
    all_missed = []
    
    for sheet_name in book.sheetnames:
        if sheet_name in [ROSTER_SHEET_NAME, MISSED_HITS_SHEET_NAME]:
            continue
        
        try:
            df = pd.read_excel(EXCEL_FILE, sheet_name=sheet_name)
            
            if df['War Complete'].iloc[0] != 'Yes':
                continue
            
            if 'Is Missed' not in df.columns:
                df['Is Missed'] = df.apply(
                    lambda row: 'Yes' if (row.get('Stars', 0) == 0 and row.get('Destruction %', 0) == 0 and row.get('Attack Number', 0) > 0) else 'No',
                    axis=1
                )
            
            missed_df = df[df['Is Missed'] == 'Yes'].copy()
            
            if not missed_df.empty:
                missed_df = missed_df[['War ID', 'War End Time', 'Player Name', 'Player Tag', 
                                      'Town Hall', 'Attack Number', 'Stars', 'Destruction %']]
                all_missed.append(missed_df)
        
        except Exception as e:
            print(f"Error reading sheet {sheet_name} for missed hits: {e}")
            continue
    
    if not all_missed:
        print("No missed hits found")
        return
    
    missed_df = pd.concat(all_missed, ignore_index=True)
    missed_df = missed_df.sort_values(['War End Time', 'Player Name'], ascending=[False, True])
    
    if MISSED_HITS_SHEET_NAME in book.sheetnames:
        del book[MISSED_HITS_SHEET_NAME]
    
    ws = book.create_sheet(MISSED_HITS_SHEET_NAME, 1)
    
    for r_idx, r in enumerate(dataframe_to_rows(missed_df, index=False, header=True), 1):
        ws.append(r)
        
        if r_idx == 1:
            for cell in ws[r_idx]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="DC3545", end_color="DC3545", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
    
    for column in ws.columns:
        max_length = 0
        column = list(column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width
    
    book.save(EXCEL_FILE)
    print(f"✓ Updated missed hits sheet with {len(missed_df)} missed attacks")

def update_roster_sheet():
    """Create/update a master roster sheet with all players"""
    if not os.path.exists(EXCEL_FILE):
        print("No war data file found")
        return
    
    book = load_workbook(EXCEL_FILE)
    all_players = {}
    
    for sheet_name in book.sheetnames:
        if sheet_name in [ROSTER_SHEET_NAME, MISSED_HITS_SHEET_NAME]:
            continue
        
        try:
            df = pd.read_excel(EXCEL_FILE, sheet_name=sheet_name)
            
            for _, row in df.iterrows():
                player_tag = row['Player Tag']
                if player_tag not in all_players:
                    all_players[player_tag] = {
                        'Player Tag': player_tag,
                        'Player Name': row['Player Name'],
                        'Last Seen TH': row.get('Town Hall', 0),
                        'Total Wars': 0,
                        'Wars Participated': 0
                    }
                
                if row.get('Attack Number', 0) > 0:
                    all_players[player_tag]['Wars Participated'] += 1
                
                current_th = row.get('Town Hall', 0)
                if current_th > all_players[player_tag]['Last Seen TH']:
                    all_players[player_tag]['Last Seen TH'] = current_th
        except Exception as e:
            print(f"Error reading sheet {sheet_name}: {e}")
            continue
    
    total_wars = len([s for s in book.sheetnames if s not in [ROSTER_SHEET_NAME, MISSED_HITS_SHEET_NAME]])
    
    for player in all_players.values():
        player['Total Wars'] = total_wars
    
    roster_df = pd.DataFrame(list(all_players.values()))
    roster_df = roster_df.sort_values('Player Name')
    
    if ROSTER_SHEET_NAME in book.sheetnames:
        del book[ROSTER_SHEET_NAME]
    
    ws = book.create_sheet(ROSTER_SHEET_NAME, 0)
    
    for r_idx, r in enumerate(dataframe_to_rows(roster_df, index=False, header=True), 1):
        ws.append(r)
        
        if r_idx == 1:
            for cell in ws[r_idx]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="2E75B5", end_color="2E75B5", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
    
    for column in ws.columns:
        max_length = 0
        column = list(column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width
    
    book.save(EXCEL_FILE)
    print(f"✓ Updated roster sheet with {len(all_players)} players")

def calculate_leaderboard(days_filter=None):
    """Calculate leaderboard statistics from completed wars"""
    if not os.path.exists(EXCEL_FILE):
        print("No war data found")
        return None
    
    book = load_workbook(EXCEL_FILE, read_only=True)
    all_data = []
    
    cutoff_date = None
    if days_filter:
        from datetime import timedelta, timezone
        cutoff_date = datetime.now(timezone.utc) - timedelta(days=days_filter)
    
    for sheet_name in book.sheetnames:
        if sheet_name in [ROSTER_SHEET_NAME, MISSED_HITS_SHEET_NAME]:
            continue
        
        try:
            df = pd.read_excel(EXCEL_FILE, sheet_name=sheet_name)
            
            if df['War Complete'].iloc[0] != 'Yes':
                continue
            
            if 'Is Missed' not in df.columns:
                df['Is Missed'] = df.apply(
                    lambda row: 'Yes' if (row.get('Stars', 0) == 0 and row.get('Destruction %', 0) == 0 and row.get('Attack Number', 0) > 0) else 'No',
                    axis=1
                )
            
            if cutoff_date and 'War End Time' in df.columns:
                df['War End Time'] = pd.to_datetime(df['War End Time'], utc=True)
                df = df[df['War End Time'] >= cutoff_date]
            
            if not df.empty:
                all_data.append(df)
        except Exception as e:
            print(f"Error processing sheet {sheet_name}: {e}")
            continue
    
    book.close()
    
    if not all_data:
        print("No completed war data found matching the filter")
        return None
    
    # Combine all data BEFORE filtering
    all_attacks_df = pd.concat(all_data, ignore_index=True)
    
    # Get all players who participated in wars (from roster, not just attacks)
    all_war_participants_full = all_attacks_df[['Player Name', 'Player Tag', 'War ID']].drop_duplicates()
    war_participation = all_war_participants_full.groupby(['Player Name', 'Player Tag']).agg({
        'War ID': 'nunique'
    }).reset_index()
    war_participation.columns = ['Player Name', 'Player Tag', 'Total Wars']
    
    # Calculate missed hits: count rows where Is Missed = Yes OR Attack Number = 0 (no attacks at all)
    missed_attacks = all_attacks_df[
        (all_attacks_df['Is Missed'] == 'Yes') | (all_attacks_df['Attack Number'] == 0)
    ].copy()
    
    # Calculate missed hits: count rows where Is Missed = Yes OR Attack Number = 0 (no attacks at all)
    # For Attack Number = 0, they get ONE row but missed 2 attacks, so we need to count properly
    def count_missed(group):
        if (group['Attack Number'] == 0).any():
            # Player didn't attack at all - they missed all expected attacks (usually 2)
            return 2  # Standard attacks per war
        else:
            # Count individual missed attacks
            return (group['Is Missed'] == 'Yes').sum()
    
    missed_stats = missed_attacks.groupby(['Player Name', 'Player Tag']).apply(count_missed).reset_index(name='Missed Hits')
    
    # Now filter for valid attacks (exclude loot and missed)
    combined_df = all_attacks_df[(all_attacks_df['Is Loot Hit'] != 'Yes') & (all_attacks_df['Is Missed'] != 'Yes')]
    combined_df = combined_df[combined_df['Attack Number'] > 0]
    
    if combined_df.empty:
        print("No valid attack data found after filtering")
        return None
    
    # Get all players who participated (including those who didn't attack)
    all_war_participants = combined_df[['Player Name', 'Player Tag', 'War ID']].drop_duplicates()
    
    # Merge with full war participation to include people who didn't attack at all
    all_war_participants = war_participation.copy()
    
    # Calculate attack statistics (only for players who attacked - excluding missed and loot)
    attack_stats = combined_df.groupby(['Player Name', 'Player Tag']).agg({
        'Stars': 'sum',
        'Is Triple': lambda x: (x == 'Yes').sum(),
        'Attack Number': 'count'
    }).reset_index()
    
    attack_stats.columns = ['Player Name', 'Player Tag', 'Total Stars', 
                           'Three Stars', 'Total Attacks']
    
    # Merge with war participation (which includes everyone in the war)
    player_stats = war_participation.merge(
        attack_stats, 
        on=['Player Name', 'Player Tag'], 
        how='left'
    )
    
    player_stats['Total Stars'] = player_stats['Total Stars'].fillna(0).astype(int)
    player_stats['Three Stars'] = player_stats['Three Stars'].fillna(0).astype(int)
    player_stats['Total Attacks'] = player_stats['Total Attacks'].fillna(0).astype(int)
    
    player_stats['3 Star Rate'] = player_stats.apply(
        lambda row: f"{(row['Three Stars'] / row['Total Attacks'] * 100):.1f}%" 
        if row['Total Attacks'] > 0 else "0.0%",
        axis=1
    )
    
    player_stats['Avg Stars Per Attack'] = player_stats.apply(
        lambda row: round(row['Total Stars'] / row['Total Attacks'], 2)
        if row['Total Attacks'] > 0 else 0.0,
        axis=1
    )
    
    print(f"\nDebug - Sample calculations:")
    if not player_stats.empty:
        sample = player_stats.head(3)
        for _, row in sample.iterrows():
            print(f"{row['Player Name']}: {row['Total Stars']} stars / {row['Total Attacks']} attacks = {row['Avg Stars Per Attack']}")
    
    # Calculate missed hits - need to check all war data, not just valid attacks
    # Merge missed hits with player stats
    player_stats = player_stats.merge(
        missed_stats,
        on=['Player Name', 'Player Tag'],
        how='left'
    )
    
    # Fill NaN with 0 for players with no missed hits
    player_stats['Missed Hits'] = player_stats['Missed Hits'].fillna(0).astype(int)
    
    player_stats = player_stats[['Player Name', 'Player Tag', '3 Star Rate', 
                                'Avg Stars Per Attack', 'Total Wars', 'Total Stars', 'Missed Hits']]
    
    # Sort by 3 Star Rate (descending), then by Missed Hits (ascending - fewer is better), then by Total Wars
    player_stats['_sort_rate'] = player_stats['3 Star Rate'].str.rstrip('%').astype(float)
    player_stats = player_stats.sort_values(['_sort_rate', 'Missed Hits', 'Total Wars'], 
                                           ascending=[False, True, False])
    player_stats = player_stats.drop(columns=['_sort_rate'])
    
    return player_stats

def save_leaderboard_json(leaderboard_data, filename=LEADERBOARD_FILE):
    """Save leaderboard data as JSON for the website"""
    if leaderboard_data is None:
        return
    
    leaderboard_list = leaderboard_data.to_dict('records')
    
    data = {
        'last_updated': datetime.now().isoformat(),
        'players': leaderboard_list
    }
    
    with open(filename, 'w') as f:
        json.dump(data, f, indent=2)
    
    print(f"✓ Saved leaderboard to {filename}")

def send_discord_war_report(war_data, war):
    """Send war completion report to Discord"""
    if not DISCORD_WEBHOOK_URL:
        return
    
    try:
        clan = war.get('clan', {})
        opponent = war.get('opponent', {})
        
        clan_name = clan.get('name', 'Unknown')
        clan_stars = clan.get('stars', 0)
        clan_destruction = clan.get('destructionPercentage', 0)
        
        opponent_name = opponent.get('name', 'Unknown')
        opponent_stars = opponent.get('stars', 0)
        opponent_destruction = opponent.get('destructionPercentage', 0)
        
        result = "🏆 VICTORY!" if clan_stars > opponent_stars else ("💔 DEFEAT" if clan_stars < opponent_stars else "🤝 TIE")
        color = 0x10b981 if clan_stars > opponent_stars else (0xef4444 if clan_stars < opponent_stars else 0xf59e0b)
        
        df = pd.DataFrame(war_data['war_details'])
        df = df[df['Attack Number'] > 0]
        
        top_stars = df.nlargest(3, 'Stars')[['Player Name', 'Stars', 'Destruction %']]
        top_performers = "\n".join([f"⭐ **{row['Player Name']}**: {row['Stars']}⭐ ({row['Destruction %']:.1f}%)" 
                                    for _, row in top_stars.iterrows()])
        
        missed_df = df[df['Is Missed'] == 'Yes']
        missed_count = len(missed_df)
        
        embed = {
            "title": f"{result}",
            "description": f"**{clan_name}** vs **{opponent_name}**",
            "color": color,
            "fields": [
                {
                    "name": "📊 Final Score",
                    "value": f"**{clan_stars}** ⭐ ({clan_destruction:.1f}%) - {opponent_stars} ⭐ ({opponent_destruction:.1f}%)",
                    "inline": False
                },
                {
                    "name": "🌟 Top Performers",
                    "value": top_performers if not top_stars.empty else "No data",
                    "inline": False
                },
                {
                    "name": "❌ Missed Attacks",
                    "value": f"{missed_count} attack(s) missed" if missed_count > 0 else "✅ No missed attacks!",
                    "inline": False
                }
            ],
            "timestamp": datetime.now().isoformat(),
            "footer": {
                "text": "100% Turtle War Tracker"
            }
        }
        
        payload = {
            "embeds": [embed]
        }
        
        response = requests.post(DISCORD_WEBHOOK_URL, json=payload)
        response.raise_for_status()
        print("✓ Sent war report to Discord")
        
    except Exception as e:
        print(f"Failed to send Discord report: {e}")

def main():
    print("="*60)
    print("CLASH OF CLANS WAR TRACKER")
    print("="*60)
    print(f"Clan: {CLAN_TAG}")
    print(f"Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("="*60)
    
    print("\nFetching current war...")
    war = fetch_current_war()
    
    if not war:
        print("Failed to fetch war data")
        return
    
    state = war.get('state', 'notInWar')
    print(f"War State: {state}")
    
    if state == 'notInWar':
        print("Clan is not currently in war")
    else:
        if is_cwl_war(war):
            print("This is a CWL war - skipping (CWL wars are excluded)")
        else:
            war_id = get_war_id(war)
            print(f"War ID: {war_id}")
            
            war_data = process_war(war)
            
            if war_data:
                war_id = war_data['war_id']
                exists, is_complete = war_already_saved(war_id)
                
                # Don't overwrite completed wars
                if is_complete:
                    print(f"War {war_id} is already complete - skipping update")
                elif war_data['is_ended']:
                    print("War has ENDED - saving complete war data...")
                    save_war_to_excel(war_data)
                    update_roster_sheet()
                    update_missed_hits_sheet()
                    send_discord_war_report(war_data, war)
                else:
                    print("War is IN PROGRESS - saving current state...")
                    save_war_to_excel(war_data)
                    update_roster_sheet()
                    update_missed_hits_sheet()
            else:
                print("Error processing war data")
    
    print("\n" + "="*60)
    print("UPDATING LEADERBOARDS (Completed Wars Only)")
    print("="*60)
    
    periods = [
        (None, 'All Time'),
        (7, 'Last 7 Days'),
        (30, 'Last 30 Days'),
        (90, 'Last 3 Months')
    ]
    
    for days, label in periods:
        print(f"\n{label}:")
        leaderboard = calculate_leaderboard(days)
        
        if leaderboard is not None and not leaderboard.empty:
            suffix = f"_{days}d" if days else "_all"
            json_filename = LEADERBOARD_FILE.replace('.json', f'{suffix}.json')
            save_leaderboard_json(leaderboard, json_filename)
            
            print(leaderboard.head(5).to_string(index=False))
        else:
            print("  No completed war data available")
    
    print("\n" + "="*60)
    print("DONE!")
    print("="*60)

if __name__ == "__main__":
    main()